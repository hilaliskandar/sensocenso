"""Etapa 11a: tabelas públicas reprodutíveis do diagnóstico TIC–TIM.

Gera as tabelas centrais especificadas no plano visual ABNT diretamente das
bases produzidas pelo pipeline. A etapa não introduz classificação, escore ou
correção retrospectiva para reproduzir a edição histórica.
"""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .paths import resolve_paths
from .proveniencia import registrar_arquivo, registrar_evento


REFERENCIA_HISTORICA_CONVERGENCIA_P75 = 1255


def _crescimento(v0: pd.Series, v1: pd.Series) -> pd.Series:
    v0 = pd.to_numeric(v0, errors="coerce")
    v1 = pd.to_numeric(v1, errors="coerce")
    return ((v1 / v0) - 1.0).where(v0.gt(0)) * 100.0


def _pivot_ano(df: pd.DataFrame, coluna: str) -> pd.DataFrame:
    tab = df.pivot(index="codigo_ibge", columns="ano", values=coluna)
    tab.columns = [f"{coluna}_{int(ano)}" for ano in tab.columns]
    return tab


def construir_t01(longitudinal: pd.DataFrame) -> pd.DataFrame:
    base = longitudinal.copy()
    base["codigo_ibge"] = base["codigo_ibge"].astype("string")
    if set(pd.to_numeric(base["ano"], errors="coerce").dropna().astype(int)) != {2000, 2010, 2022}:
        raise ValueError("T01 exige os anos 2000, 2010 e 2022.")
    if base.duplicated(["codigo_ibge", "ano"]).any():
        raise ValueError("T01 recebeu mais de uma linha por município/ano.")

    municipio = (
        base.loc[pd.to_numeric(base["ano"], errors="coerce").eq(2022), ["codigo_ibge", "municipio_config"]]
        .drop_duplicates("codigo_ibge")
        .set_index("codigo_ibge")
    )
    partes = [municipio]
    for coluna in ("pop_total_harmonizada", "pop_0_14", "pop_60_mais", "razao_envelhecimento"):
        partes.append(_pivot_ano(base, coluna))
    out = pd.concat(partes, axis=1)

    for ano in (2000, 2022):
        total = out[f"pop_total_harmonizada_{ano}"]
        out[f"pct_0_14_{ano}"] = 100.0 * out[f"pop_0_14_{ano}"] / total
        out[f"pct_60_mais_{ano}"] = 100.0 * out[f"pop_60_mais_{ano}"] / total
    out["crescimento_pop_2000_2010_pct"] = _crescimento(
        out["pop_total_harmonizada_2000"], out["pop_total_harmonizada_2010"]
    )
    out["crescimento_pop_2010_2022_pct"] = _crescimento(
        out["pop_total_harmonizada_2010"], out["pop_total_harmonizada_2022"]
    )

    colunas = [
        "municipio_config",
        "pop_total_harmonizada_2000",
        "pop_total_harmonizada_2010",
        "pop_total_harmonizada_2022",
        "crescimento_pop_2000_2010_pct",
        "crescimento_pop_2010_2022_pct",
        "pct_0_14_2000",
        "pct_0_14_2022",
        "pct_60_mais_2000",
        "pct_60_mais_2022",
        "razao_envelhecimento_2022",
    ]
    out = out[colunas].rename(columns={"municipio_config": "municipio"}).reset_index()
    if len(out) != 30 or out["codigo_ibge"].nunique() != 30:
        raise AssertionError("T01 não fechou em 30 municípios.")
    return out.sort_values("municipio").reset_index(drop=True)


def construir_t06(familias: pd.DataFrame) -> pd.DataFrame:
    base = familias.copy()
    base["codigo_ibge"] = base["codigo_ibge"].astype("string")
    integrado = base.loc[base["FLAG_UNIVERSO_INTEGRADO"].astype(bool)].copy()
    if len(integrado) != 8073:
        raise AssertionError(f"T06 exige 8.073 setores integrados; obtidos={len(integrado)}")
    integrado["POP_TOTAL"] = pd.to_numeric(integrado["POP_TOTAL"], errors="coerce")
    integrado["DPPO"] = pd.to_numeric(integrado["DPPO"], errors="coerce")
    # CONVERGENCIA_3_OU_4 é triestado na etapa 07: 1=convergente,
    # 0=não convergente, NA=indeterminado por cobertura. Para a T06, apenas
    # o valor explicitamente igual a 1 entra no numerador; NA permanece fora
    # do numerador, mas o setor continua pertencendo ao denominador integrado.
    integrado["conv"] = (
        pd.to_numeric(integrado["CONVERGENCIA_3_OU_4"], errors="coerce")
        .eq(1)
        .fillna(False)
        .astype(bool)
    )
    integrado["pop_conv"] = integrado["POP_TOTAL"].where(integrado["conv"], 0)
    integrado["dppo_conv"] = integrado["DPPO"].where(integrado["conv"], 0)
    integrado["setor_conv"] = integrado["conv"].astype("int64")

    out = integrado.groupby(["codigo_ibge", "municipio"], as_index=False).agg(
        setores_integrados=("codigo_setor", "size"),
        setores_convergentes=("setor_conv", "sum"),
        populacao_integrada=("POP_TOTAL", "sum"),
        populacao_convergente=("pop_conv", "sum"),
        dppo_integrados=("DPPO", "sum"),
        dppo_convergentes=("dppo_conv", "sum"),
    )
    out["pct_setores_convergentes"] = 100.0 * out["setores_convergentes"] / out["setores_integrados"]
    out["pct_populacao_convergente"] = (
        100.0 * out["populacao_convergente"] / out["populacao_integrada"]
    ).where(out["populacao_integrada"].gt(0))
    out["pct_dppo_convergentes"] = (
        100.0 * out["dppo_convergentes"] / out["dppo_integrados"]
    ).where(out["dppo_integrados"].gt(0))
    if len(out) != 30:
        raise AssertionError("T06 não fechou em 30 municípios.")
    return out.sort_values("municipio").reset_index(drop=True)


def _pct_familia(integrado: pd.DataFrame, coluna: str) -> pd.Series:
    flag = pd.to_numeric(integrado[coluna], errors="coerce")
    numerador = flag.eq(1).groupby(integrado["codigo_ibge"]).sum()
    denominador = flag.notna().groupby(integrado["codigo_ibge"]).sum()
    return (100.0 * numerador / denominador).where(denominador.gt(0))


def construir_t10(
    longitudinal: pd.DataFrame,
    domicilios: pd.DataFrame,
    renovacao: pd.DataFrame,
    sintese: pd.DataFrame,
    distributivas: pd.DataFrame,
    familias: pd.DataFrame,
) -> pd.DataFrame:
    long22 = longitudinal.loc[pd.to_numeric(longitudinal["ano"], errors="coerce").eq(2022)].copy()
    long10 = longitudinal.loc[pd.to_numeric(longitudinal["ano"], errors="coerce").eq(2010)].copy()
    dom22 = domicilios.loc[pd.to_numeric(domicilios["ano"], errors="coerce").eq(2022)].copy()
    dom10 = domicilios.loc[pd.to_numeric(domicilios["ano"], errors="coerce").eq(2010)].copy()

    for df in (long22, long10, dom22, dom10, renovacao, sintese, distributivas, familias):
        df["codigo_ibge"] = df["codigo_ibge"].astype("string")

    base = long22[["codigo_ibge", "municipio_config", "pop_total_harmonizada", "razao_envelhecimento"]].rename(
        columns={"municipio_config": "municipio", "pop_total_harmonizada": "populacao_2022", "razao_envelhecimento": "razao_envelhecimento_2022"}
    ).set_index("codigo_ibge")
    pop10 = long10.set_index("codigo_ibge")["pop_total_harmonizada"]
    base["crescimento_pop_2010_2022_pct"] = _crescimento(pop10.reindex(base.index), base["populacao_2022"])

    dpo22 = dom22.set_index("codigo_ibge")["dpo"]
    dpo10 = dom10.set_index("codigo_ibge")["dpo"]
    base["crescimento_dpo_2010_2022_pct"] = _crescimento(dpo10.reindex(base.index), dpo22.reindex(base.index))
    base["pct_unipessoais_2022"] = dom22.set_index("codigo_ibge")["pct_unipessoais"].reindex(base.index) * 100.0
    base["cwr_0_4_por_1000_m1549"] = renovacao.set_index("codigo_ibge")["cwr_0_4_por_1000_m1549"].reindex(base.index)

    integrado = familias.loc[familias["FLAG_UNIVERSO_INTEGRADO"].astype(bool)].copy()
    for coluna in ("F1", "F2", "F3", "F4"):
        base[f"pct_setores_{coluna.lower()}"] = _pct_familia(integrado, coluna).reindex(base.index)

    s = sintese.set_index("codigo_ibge")
    base["pct_setores_convergencia_3ou4"] = s["pct_setores_convergencia_3ou4"].reindex(base.index)
    base["gravidade_fisico_urbana"] = s["gravidade_fisico_urbana"].reindex(base.index)
    d = distributivas.set_index("codigo_ibge")
    base["pct_preta_parda_urbano"] = 100.0 * d["pct_preta_parda_urbano"].reindex(base.index)
    base["pct_pop_urbana_em_fcu"] = 100.0 * d["pct_pop_urbana_em_fcu"].reindex(base.index)

    out = base.reset_index()
    if len(out) != 30 or out["codigo_ibge"].nunique() != 30:
        raise AssertionError("T10 não fechou em 30 municípios.")
    return out.sort_values("municipio").reset_index(drop=True)


def _formatar_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            letra = get_column_letter(col)
            largura = max(10, min(28, max(len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1)) + 2))
            ws.column_dimensions[letra].width = largura
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def executar(raiz: Path) -> None:
    raiz = raiz.resolve()
    paths = resolve_paths(raiz)
    paths.create()
    manifesto = paths.manifests / "execucao.jsonl"
    arquivos = {
        "longitudinal": paths.processed / "municipal" / "base_longitudinal_2000_2010_2022.parquet",
        "domicilios": paths.processed / "municipal" / "base_domiciliar_2000_2010_2022.parquet",
        "renovacao": paths.processed / "municipal" / "base_renovacao_demografica_2022.parquet",
        "sintese": paths.processed / "municipal" / "base_sintese_municipal_2022.parquet",
        "distributivas": paths.processed / "municipal" / "base_camadas_distributivas_2022.parquet",
        "familias": paths.processed / "setorial" / "base_familias_analiticas_p75.parquet",
    }
    for nome, path in arquivos.items():
        if not path.exists():
            raise FileNotFoundError(f"Pré-requisito 11a ausente ({nome}): {path}")

    dados = {nome: pd.read_parquet(path) for nome, path in arquivos.items()}
    t01 = construir_t01(dados["longitudinal"])
    t06 = construir_t06(dados["familias"])
    t10 = construir_t10(
        dados["longitudinal"], dados["domicilios"], dados["renovacao"], dados["sintese"], dados["distributivas"], dados["familias"]
    )

    paths.tables.mkdir(parents=True, exist_ok=True)
    saidas = []
    for nome, df in (("T01_populacao_transformacao_demografica", t01), ("T06_populacao_domicilios_areas_combinadas", t06), ("T10_panorama_comparativo_30_municipios", t10)):
        csv_path = paths.tables / f"{nome}.csv"
        df.to_csv(csv_path, index=False, encoding="utf-8")
        registrar_arquivo(manifesto, csv_path, origem="Etapa 11a — tabela pública")
        saidas.append(str(csv_path.relative_to(paths.data_root)))

    xlsx_path = paths.tables / "TIC_TIM_TABELAS_PUBLICAS_REPRODUTIVEIS_11a.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        t01.to_excel(writer, sheet_name="T01", index=False)
        t06.to_excel(writer, sheet_name="T06", index=False)
        t10.to_excel(writer, sheet_name="T10", index=False)
    _formatar_xlsx(xlsx_path)
    registrar_arquivo(manifesto, xlsx_path, origem="Etapa 11a — conjunto de tabelas públicas")
    saidas.append(str(xlsx_path.relative_to(paths.data_root)))

    convergencia_corrente = int(t06["setores_convergentes"].sum())
    qa = {
        "status": "OK_COM_DERIVA_EDICAO" if convergencia_corrente != REFERENCIA_HISTORICA_CONVERGENCIA_P75 else "OK_REPRODUCAO_REFERENCIA",
        "etapa": "11a",
        "tabelas": {
            "T01": {"linhas": int(len(t01)), "municipios": int(t01["codigo_ibge"].nunique())},
            "T06": {"linhas": int(len(t06)), "municipios": int(t06["codigo_ibge"].nunique())},
            "T10": {"linhas": int(len(t10)), "municipios": int(t10["codigo_ibge"].nunique())},
        },
        "universo_integrado_t06": int(t06["setores_integrados"].sum()),
        "setores_convergentes_corrente": convergencia_corrente,
        "setores_convergentes_referencia_historica": REFERENCIA_HISTORICA_CONVERGENCIA_P75,
        "delta_convergencia": convergencia_corrente - REFERENCIA_HISTORICA_CONVERGENCIA_P75,
        "regra": "tabelas derivadas exclusivamente das bases do pipeline; referências históricas são QA e não alvos de calibração",
        "saidas": saidas,
    }
    qa_path = paths.qa / "etapa11a_tabelas.json"
    qa_path.write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    registrar_arquivo(manifesto, qa_path, origem="Etapa 11a — QA tabelas")
    registrar_evento(manifesto, {"tipo": "etapa", "etapa": "11a", "status": qa["status"], "tabelas": 3, "setores_convergentes": convergencia_corrente})
    print(json.dumps(qa, ensure_ascii=False, indent=2))