from __future__ import annotations

import io
import json
import re
import unicodedata
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from .paths import resolve_paths
from .proveniencia import registrar_arquivo, registrar_evento


VAR_ENTORNO_RE = re.compile(r"\bV0(?:50|52|54)\d{2}\b", re.I)
UNIVERSO_POR_PREFIXO = {"V050": "domicilios", "V052": "moradores", "V054": "faces"}
ATRIBUTOS = (
    "bueiro_boca_de_lobo",
    "calcada",
    "pavimentacao",
    "iluminacao_publica",
    "arborizacao",
    "rampa_cadeirante",
    "obstaculo_calcada",
    "ponto_onibus",
    "infraestrutura_cicloviaria",
)


def _normalizar(valor: object) -> str:
    texto = " ".join(str(valor or "").replace("\n", " ").split()).strip()
    texto = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in texto if not unicodedata.combining(c)).casefold()


def _texto_original(valores: list[object]) -> str:
    return " | ".join(
        " ".join(str(v or "").replace("\n", " ").split()).strip()
        for v in valores
        if v not in (None, "")
    )


def _atributos_da_linha(norm: str) -> list[str]:
    out: list[str] = []
    if "bueiro" in norm or "boca de lobo" in norm or "boca-de-lobo" in norm:
        out.append("bueiro_boca_de_lobo")
    if "calcada" in norm and "obstaculo" not in norm:
        out.append("calcada")
    if "paviment" in norm:
        out.append("pavimentacao")
    if "iluminacao publica" in norm:
        out.append("iluminacao_publica")
    if "arborizacao" in norm:
        out.append("arborizacao")
    if "rampa para cadeirante" in norm:
        out.append("rampa_cadeirante")
    if "obstaculo" in norm and "calcada" in norm:
        out.append("obstaculo_calcada")
    if "ponto de onibus" in norm or "ponto de onibus / van" in norm or "parada de onibus" in norm:
        out.append("ponto_onibus")
    if "via sinalizada para bicicleta" in norm or "via sinalizada para bicicletas" in norm:
        out.append("infraestrutura_cicloviaria")
    return out


def _categoria(texto_norm: str, atributo: str) -> str | None:
    if atributo == "arborizacao":
        if "saltado" in texto_norm:
            return "nao_declarado"
        if "sem arvores" in texto_norm:
            return "nao"
        if any(x in texto_norm for x in ("de 1 a 2 arvores", "de 3 a 4 arvores", "5 ou mais arvores")):
            return "sim"
        return None
    if "nao declarado" in texto_norm or "nao-declarado" in texto_norm:
        return "nao_declarado"
    if re.search(r"(?:^|\W)nao(?:\W|$)", texto_norm) or "sem " in texto_norm:
        return "nao"
    if re.search(r"(?:^|\W)sim(?:\W|$)", texto_norm) or "com " in texto_norm:
        return "sim"
    return None


def _inspecionar_workbook(wb, origem: str) -> dict:
    achados: dict[str, list[dict]] = {k: [] for k in ATRIBUTOS}
    for ws in wb.worksheets:
        linhas = [list(row) for row in ws.iter_rows(values_only=True)]
        for i, valores in enumerate(linhas):
            original = _texto_original(valores)
            norm = _normalizar(original)
            atributos = _atributos_da_linha(norm)
            if not atributos:
                continue
            contexto_linhas = linhas[max(0, i - 1) : min(len(linhas), i + 2)]
            contexto = " || ".join(_texto_original(v) for v in contexto_linhas)
            codigos = sorted(set(VAR_ENTORNO_RE.findall(original.upper())))
            if not codigos:
                codigos = sorted(set(VAR_ENTORNO_RE.findall(contexto.upper())))
            for atributo in atributos:
                achados[atributo].append(
                    {
                        "origem": origem,
                        "planilha": ws.title,
                        "linha": i + 1,
                        "texto": original,
                        "contexto": contexto if contexto != original else None,
                        "categoria": _categoria(norm, atributo),
                        "codigos": codigos,
                    }
                )
    return achados


def _mesclar_achados(destino: dict, origem: dict) -> None:
    for atributo, itens in origem.items():
        destino.setdefault(atributo, []).extend(itens)


def inspecionar_dicionario(path: Path) -> dict:
    achados: dict[str, list[dict]] = {k: [] for k in ATRIBUTOS}
    if path.suffix.casefold() == ".zip":
        with zipfile.ZipFile(path) as zf:
            membros = [m for m in zf.namelist() if m.casefold().endswith(".xlsx")]
            if not membros:
                raise ValueError(f"ZIP de documentação sem XLSX: {path}")
            for membro in membros:
                wb = load_workbook(io.BytesIO(zf.read(membro)), read_only=True, data_only=True)
                try:
                    _mesclar_achados(achados, _inspecionar_workbook(wb, membro))
                finally:
                    wb.close()
        return achados

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return _inspecionar_workbook(wb, path.name)
    finally:
        wb.close()


def resumir_codigos(achados: dict[str, list[dict]]) -> dict:
    resumo: dict[str, dict] = {}
    for atributo, linhas in achados.items():
        por_universo = {
            u: {"sim": set(), "nao": set(), "nao_declarado": set(), "sem_categoria": set()}
            for u in UNIVERSO_POR_PREFIXO.values()
        }
        for item in linhas:
            categoria = item.get("categoria") or "sem_categoria"
            for codigo in item.get("codigos", []):
                codigo = str(codigo).upper()
                universo = UNIVERSO_POR_PREFIXO.get(codigo[:4])
                if universo:
                    por_universo[universo][categoria].add(codigo)
        resumo[atributo] = {
            universo: {cat: sorted(vals) for cat, vals in cats.items()}
            for universo, cats in por_universo.items()
        }
    return resumo


def _resolver(resumo: dict, cabecalhos: dict[str, set[str]]) -> dict:
    resolvido: dict[str, dict] = {}
    for atributo, universos in resumo.items():
        resolvido[atributo] = {}
        for universo, cats in universos.items():
            sim = list(cats.get("sim", []))
            nao = list(cats.get("nao", []))
            nd = list(cats.get("nao_declarado", []))
            sem = list(cats.get("sem_categoria", []))
            todos = sim + nao + nd
            valido = bool(sim and nao and not sem and all(c in cabecalhos[universo] for c in todos))
            resolvido[atributo][universo] = {
                "sim": sim,
                "nao": nao,
                "nao_declarado": nd,
                "confirmados_no_cabecalho": bool(valido),
            }
    return resolvido


def executar(raiz: Path) -> None:
    raiz = raiz.resolve()
    paths = resolve_paths(raiz)
    paths.create()
    manifesto = paths.manifests / "execucao.jsonl"

    qa05b_path = paths.qa / "etapa05b_inspecao_fontes_isau.json"
    if not qa05b_path.exists():
        raise FileNotFoundError(f"Pré-requisito 06a ausente: {qa05b_path}")
    qa05b = json.loads(qa05b_path.read_text(encoding="utf-8"))
    nome_resolvedor = qa05b.get("dicionario_resolvedor_bueiro")
    if not nome_resolvedor:
        raise ValueError("05b não registrou dicionário oficial resolvedor de bueiro.")

    cabecalhos: dict[str, set[str]] = {}
    for universo in ("domicilios", "moradores", "faces"):
        cabecalhos[universo] = {
            c
            for info in qa05b["inspecao_entorno"][universo]["csvs"]
            for c in info["colunas"]
        }

    raw_doc = paths.raw / "ibge" / "censo2022" / "isau" / "documentacao"
    candidatos = sorted(list(raw_doc.glob("*.xlsx")) + list(raw_doc.glob("*.zip")))
    if not candidatos:
        raise FileNotFoundError(
            "Nenhum dicionário XLSX/ZIP materializado por 05b. A etapa 06a não presume nome ou códigos."
        )

    resultados = []
    resolucao_final = None
    for arquivo in candidatos:
        try:
            achados = inspecionar_dicionario(arquivo)
            n_total = sum(len(v) for v in achados.values())
            if n_total:
                resumo = resumir_codigos(achados)
                resolucao = _resolver(resumo, cabecalhos)
                resultados.append(
                    {
                        "arquivo": arquivo.name,
                        "n_achados": n_total,
                        "achados": achados,
                        "resumo_codigos": resumo,
                        "resolucao": resolucao,
                    }
                )
                if all(
                    resolucao[a][u]["confirmados_no_cabecalho"]
                    for a in ATRIBUTOS
                    for u in ("domicilios", "moradores", "faces")
                ):
                    resolucao_final = resolucao
        except Exception as exc:
            resultados.append({"arquivo": arquivo.name, "erro": f"{type(exc).__name__}: {exc}"})

    if resolucao_final is None:
        status = "DIAGNOSTICO_SEMANTICO_PENDENTE"
    else:
        status = "RESOLVIDO_ENTORNO"

    qa = {
        "status": status,
        "etapa": "06a",
        "objetivo": "descobrir e confirmar nos cabeçalhos oficiais os códigos dos atributos de entorno",
        "atributos_nucleares_f3": [
            "bueiro_boca_de_lobo", "calcada", "pavimentacao", "iluminacao_publica", "arborizacao"
        ],
        "atributos_complementares": [
            "rampa_cadeirante", "obstaculo_calcada", "ponto_onibus", "infraestrutura_cicloviaria"
        ],
        "dicionario_resolvedor_05b": nome_resolvedor,
        "resultados": resultados,
        "codigos_resolvidos": resolucao_final,
        "regra": (
            "Percentuais de ausência usam apenas categorias substantivas no denominador. Em variáveis binárias, "
            "denominador=Sim+Não e Não declarado é excluído. Em arborização, Sem árvores é ausência; as faixas "
            "1–2, 3–4 e 5+ árvores formam presença; Saltado é excluído do denominador."
        ),
        "regra_f3": (
            "F3 usa, no agregado final, os cinco percentuais de ausência segundo moradores; sinaliza pelo menos "
            "dois componentes no P75 e, quando P75=0, exige valor estritamente positivo."
        ),
        "proximo_gate": "06b somente se status=RESOLVIDO_ENTORNO",
    }
    destino = paths.qa / "etapa06a_gate_semantico_entorno.json"
    destino.write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    registrar_arquivo(manifesto, destino, origem="Gate semântico 06a - entorno urbano")
    registrar_evento(manifesto, {"tipo": "etapa", "etapa": "06a", "status": qa["status"]})
    print(json.dumps(qa, ensure_ascii=False, indent=2))
