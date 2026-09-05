"""Etapa 11e: manifesto visual e QA de cobertura do diagnóstico TIC–TIM.

A etapa não recalcula indicadores. Ela audita as saídas editoriais produzidas
pelas etapas 11a–11d contra o inventário visual final congelado no repositório,
verifica integridade básica dos arquivos e persiste um manifesto legível por
máquina. O quadro Q01 é um elemento editorial já produzido no Relatório
Regional v1.8; por ser síntese textual dos capítulos 2–4, é contabilizado como
artefato editorial externo ao pipeline, e não como saída numérica recalculada.
"""
from __future__ import annotations

import csv
import hashlib
import json
from dataclasses import asdict, dataclass
from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from PIL import Image

from .paths import resolve_paths
from .proveniencia import registrar_arquivo, registrar_evento


RELATORIO_Q01_ID = "1z_ZIF6v82pT_h5tai_h7JTVyimlXsbbM"
RELATORIO_Q01_TITULO = "TIC_TIM_DEM_HAB_RELATORIO_REGIONAL_v1.8.docx"


@dataclass(frozen=True)
class ElementoVisual:
    id: str
    capitulo: str
    tipo: str
    titulo: str
    escopo: str
    primario: str | None
    png: str | None = None
    svg: str | None = None
    base: str | None = None
    externo: bool = False
    nota: str = ""


def inventario_visual() -> tuple[ElementoVisual, ...]:
    base11c = "outputs/data/11c/base_cartografia_municipal_30.csv"
    return (
        ElementoVisual("T01", "2", "Tabela", "População, crescimento e transformação demográfica dos 30 municípios, 2000–2022", "municipal", "outputs/tables/T01_populacao_transformacao_demografica.csv"),
        ElementoVisual("M01", "2", "Mapa", "Crescimento populacional entre 2010 e 2022", "municipal", "outputs/maps/M01.png", "outputs/maps/M01.png", "outputs/maps/M01.svg", base11c),
        ElementoVisual("M02", "2", "Mapa", "Envelhecimento da população em 2022", "municipal", "outputs/maps/M02.png", "outputs/maps/M02.png", "outputs/maps/M02.svg", base11c),
        ElementoVisual("M03", "2", "Mapa", "Renovação geracional em 2022", "municipal", "outputs/maps/M03.png", "outputs/maps/M03.png", "outputs/maps/M03.svg", base11c, nota="CWR é proxy censitária."),
        ElementoVisual("M04", "2", "Mapa", "Renovação geracional em escala local", "setorial", "outputs/maps/M04.png", "outputs/maps/M04.png", "outputs/maps/M04.svg", "outputs/data/11d/M04_renovacao_geracional_setorial.csv"),
        ElementoVisual("G02", "2", "Gráfico", "Mudança da estrutura etária entre 2000 e 2022", "regional", "outputs/graphs/G02.png", "outputs/graphs/G02.png", "outputs/graphs/G02.svg", "outputs/data/etapa11b/G02_dados.csv"),
        ElementoVisual("G03", "2", "Gráfico", "Crescimento populacional e envelhecimento", "municipal", "outputs/graphs/G03.png", "outputs/graphs/G03.png", "outputs/graphs/G03.svg", "outputs/data/etapa11b/G03_dados.csv"),
        ElementoVisual("G04", "2", "Gráfico", "Crescimento dos domicílios e da população, 2010–2022", "municipal", "outputs/graphs/G04.png", "outputs/graphs/G04.png", "outputs/graphs/G04.svg", "outputs/data/etapa11b/G04_dados.csv"),
        ElementoVisual("G05", "2", "Gráfico", "Redução do tamanho médio dos domicílios, 2000–2022", "municipal", "outputs/graphs/G05.png", "outputs/graphs/G05.png", "outputs/graphs/G05.svg", "outputs/data/etapa11b/G05_dados.csv"),
        ElementoVisual("G06", "2", "Gráfico", "Crescimento dos domicílios unipessoais, 2000–2022", "regional/municipal", "outputs/graphs/G06.png", "outputs/graphs/G06.png", "outputs/graphs/G06.svg", "outputs/data/etapa11b/G06_municipios_dados.csv"),
        ElementoVisual("M05", "2", "Mapa", "Participação da população preta e parda na população urbana com informação válida de cor ou raça", "municipal", "outputs/maps/M05.png", "outputs/maps/M05.png", "outputs/maps/M05.svg", base11c),
        ElementoVisual("G11", "2", "Gráfico", "Participação preta e parda e gravidade físico-urbana", "municipal", "outputs/graphs/G11.png", "outputs/graphs/G11.png", "outputs/graphs/G11.svg", "outputs/data/etapa11b/G11_dados.csv", nota="Associação ecológica e descritiva."),
        ElementoVisual("M06", "3", "Mapa", "Privação sanitário-ambiental associada à moradia", "setorial", "outputs/maps/M06.png", "outputs/maps/M06.png", "outputs/maps/M06.svg", "outputs/data/11d/M06_privacao_sanitario_ambiental_setorial.csv"),
        ElementoVisual("M08", "3", "Mapa", "Áreas com necessidades habitacionais combinadas", "setorial", "outputs/maps/M08.png", "outputs/maps/M08.png", "outputs/maps/M08.svg", "outputs/data/11d/M08_necessidades_combinadas_p75.csv", nota="Edição corrente; referência histórica P75=1.255, sem calibração."),
        ElementoVisual("T06", "3", "Tabela", "População e domicílios em áreas com necessidades combinadas, por município", "municipal/setorial", "outputs/tables/T06_populacao_domicilios_areas_combinadas.csv"),
        ElementoVisual("M10", "4", "Mapa", "Domicílios com abastecimento de água fora da rede geral", "municipal", "outputs/maps/M10.png", "outputs/maps/M10.png", "outputs/maps/M10.svg", base11c),
        ElementoVisual("M11", "4", "Mapa", "Domicílios com esgotamento sanitário inadequado", "municipal", "outputs/maps/M11.png", "outputs/maps/M11.png", "outputs/maps/M11.svg", base11c),
        ElementoVisual("M12", "4", "Prancha cartográfica", "Principais carências da infraestrutura do entorno urbano", "setorial", "outputs/maps/M12.png", "outputs/maps/M12.png", "outputs/maps/M12.svg", "outputs/data/11d/M12_carencias_entorno_cinco_componentes.csv"),
        ElementoVisual("G09", "4", "Gráfico", "Composição das carências físico-sanitárias", "regional", "outputs/graphs/G09.png", "outputs/graphs/G09.png", "outputs/graphs/G09.svg", "outputs/data/etapa11b/G09_dados.csv"),
        ElementoVisual("G12", "4", "Gráfico", "Carências físico-sanitárias e carências do entorno", "municipal", "outputs/graphs/G12.png", "outputs/graphs/G12.png", "outputs/graphs/G12.svg", "outputs/data/etapa11b/G12_dados.csv", nota="Índices comparativos relativos, não percentuais."),
        ElementoVisual("Q01", "5", "Quadro", "Campos de qualificação local e evidências que os sustentam", "regional", None, externo=True, nota=f"Produzido e inserido no {RELATORIO_Q01_TITULO}, Drive ID {RELATORIO_Q01_ID}; síntese editorial dos capítulos 2–4, não programa de obras."),
        ElementoVisual("M14", "6", "Mapa", "Panorama regional das dimensões predominantes", "municipal", "outputs/maps/M14.png", "outputs/maps/M14.png", "outputs/maps/M14.svg", base11c, nota="Categorias não ordinais; empates preservados."),
        ElementoVisual("G13", "6", "Gráfico", "Crescimento domiciliar e gravidade físico-urbana nos 30 municípios", "municipal", "outputs/graphs/G13.png", "outputs/graphs/G13.png", "outputs/graphs/G13.svg", "outputs/data/etapa11b/G13_dados.csv"),
        ElementoVisual("T10", "6", "Tabela", "Panorama comparativo dos 30 municípios", "municipal", "outputs/tables/T10_panorama_comparativo_30_municipios.csv"),
        ElementoVisual("M09", "Anexo metodológico", "Mapa", "Estabilidade das áreas de necessidades combinadas sob critério mais restritivo", "setorial", "outputs/maps/M09.png", "outputs/maps/M09.png", "outputs/maps/M09.svg", "outputs/data/11d/M09_estabilidade_p75_p80.csv", nota="Edição corrente; referências históricas P75=1.255 e persistentes P80=959."),
    )


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for bloco in iter(lambda: f.read(1024 * 1024), b""):
            h.update(bloco)
    return h.hexdigest()


def _validar_png(path: Path) -> dict[str, int]:
    with Image.open(path) as im:
        im.verify()
    with Image.open(path) as im:
        largura, altura = im.size
    if largura < 1000 or altura < 600:
        raise AssertionError(f"PNG abaixo da dimensão mínima de QA: {path} ({largura}x{altura})")
    return {"largura_px": int(largura), "altura_px": int(altura)}


def _validar_svg(path: Path) -> None:
    raiz = ET.parse(path).getroot()
    if not str(raiz.tag).endswith("svg"):
        raise AssertionError(f"SVG inválido: {path}")


def _validar_csv(path: Path) -> dict[str, int]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        leitor = csv.reader(f)
        linhas = list(leitor)
    if len(linhas) < 2 or not linhas[0]:
        raise AssertionError(f"CSV vazio ou sem cabeçalho: {path}")
    return {"linhas_dados": len(linhas) - 1, "colunas": len(linhas[0])}


def _ler_qa(path: Path) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"QA de pré-requisito ausente: {path}")
    dados = json.loads(path.read_text(encoding="utf-8"))
    if not str(dados.get("status", "")).startswith("OK"):
        raise AssertionError(f"QA de pré-requisito não está OK: {path.name}: {dados.get('status')}")
    return dados


def _validar_invariantes(paths) -> dict[str, object]:
    qa11a = _ler_qa(paths.qa / "etapa11a_tabelas.json")
    qa11b = _ler_qa(paths.qa / "etapa11b_graficos.json")
    qa11c = _ler_qa(paths.qa / "etapa11c_cartografia_municipal.json")
    qa11d = _ler_qa(paths.qa / "etapa11d_cartografia_setorial.json")

    if set(qa11a.get("tabelas", {})) != {"T01", "T06", "T10"}:
        raise AssertionError("11a não contém exatamente T01, T06 e T10.")
    if int(qa11b.get("n_graficos", -1)) != 9:
        raise AssertionError("11b não contém nove gráficos.")
    if int(qa11c.get("municipios", -1)) != 30 or not qa11c.get("territorio_municipal_integral"):
        raise AssertionError("11c não preserva 30 territórios municipais integrais.")
    inv = qa11d.get("invariantes_correntes", {})
    esperados = {
        "M04_validos": 7474,
        "M06_validos": 8067,
        "M08_convergentes": 1304,
        "M09_persistentes": 1016,
        "M09_mesmo_vetor": 945,
        "M12_universo": 9087,
    }
    diverg = {k: {"observado": inv.get(k), "esperado": v} for k, v in esperados.items() if inv.get(k) != v}
    if diverg:
        raise AssertionError(f"Invariantes 11d divergiram: {diverg}")
    return {
        "11a_tabelas": sorted(qa11a["tabelas"]),
        "11b_graficos": int(qa11b["n_graficos"]),
        "11c_municipios": int(qa11c["municipios"]),
        "11c_territorio_municipal_integral": bool(qa11c["territorio_municipal_integral"]),
        "11d_invariantes": esperados,
    }


def executar(raiz: Path) -> None:
    raiz = raiz.resolve()
    paths = resolve_paths(raiz)
    paths.create()
    manifesto_execucao = paths.manifests / "execucao.jsonl"
    inventario = inventario_visual()
    ids = [e.id for e in inventario]
    if len(inventario) != 25 or len(set(ids)) != 25:
        raise AssertionError(f"Inventário visual final deve conter 25 IDs únicos; obtidos={len(inventario)}")

    integridade: list[dict[str, object]] = []
    linhas_manifesto: list[dict[str, object]] = []
    faltantes: list[str] = []
    for elemento in inventario:
        linha = asdict(elemento)
        if elemento.externo:
            linha["status"] = "EDITORIAL_EXTERNO_CONTABILIZADO"
            linhas_manifesto.append(linha)
            continue
        obrigatorios = [p for p in (elemento.primario, elemento.png, elemento.svg, elemento.base) if p]
        for rel in dict.fromkeys(obrigatorios):
            path = paths.data_root / str(rel)
            if not path.exists() or path.stat().st_size <= 0:
                faltantes.append(f"{elemento.id}:{rel}")
                continue
            registro: dict[str, object] = {
                "id": elemento.id,
                "arquivo": str(rel),
                "bytes": int(path.stat().st_size),
                "sha256": _sha256(path),
            }
            suf = path.suffix.casefold()
            if suf == ".png":
                registro.update(_validar_png(path))
            elif suf == ".svg":
                _validar_svg(path)
            elif suf == ".csv":
                registro.update(_validar_csv(path))
            integridade.append(registro)
        linha["status"] = "OK_PIPELINE" if not any(x.startswith(f"{elemento.id}:") for x in faltantes) else "FALTANTE"
        linhas_manifesto.append(linha)

    if faltantes:
        raise AssertionError(f"Cobertura visual incompleta: {faltantes}")

    xlsx = paths.tables / "TIC_TIM_TABELAS_PUBLICAS_REPRODUTIVEIS_11a.xlsx"
    if not xlsx.exists():
        raise FileNotFoundError(f"XLSX consolidado da 11a ausente: {xlsx}")
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    abas = set(wb.sheetnames)
    wb.close()
    if abas != {"T01", "T06", "T10"}:
        raise AssertionError(f"Abas inesperadas no XLSX 11a: {sorted(abas)}")
    integridade.append({"id": "11a_XLSX", "arquivo": str(xlsx.relative_to(paths.data_root)), "bytes": int(xlsx.stat().st_size), "sha256": _sha256(xlsx), "abas": "T01|T06|T10"})

    invariantes = _validar_invariantes(paths)
    manifesto_csv = paths.qa / "etapa11e_manifesto_visual.csv"
    integridade_csv = paths.qa / "etapa11e_integridade_arquivos.csv"
    qa_json = paths.qa / "etapa11e_manifesto_visual.json"

    import pandas as pd
    pd.DataFrame(linhas_manifesto).to_csv(manifesto_csv, index=False, encoding="utf-8")
    pd.DataFrame(integridade).to_csv(integridade_csv, index=False, encoding="utf-8")

    n_pipeline = sum(not e.externo for e in inventario)
    n_externos = sum(e.externo for e in inventario)
    qa = {
        "status": "OK_COBERTURA_VISUAL",
        "etapa": "11e",
        "fonte_especificacao": "docs/ETAPA11_INVENTARIO_VISUAL_FINAL.md; matriz editorial canônica, aba 12_PLANO_VISUAL_ABNT",
        "elementos_planejados": len(inventario),
        "elementos_pipeline": n_pipeline,
        "elementos_editoriais_externos": n_externos,
        "elementos_contabilizados": n_pipeline + n_externos,
        "ids": ids,
        "faltantes": [],
        "q01": {
            "status": "EDITORIAL_EXTERNO_CONTABILIZADO",
            "arquivo": RELATORIO_Q01_TITULO,
            "drive_id": RELATORIO_Q01_ID,
            "regra": "síntese textual dos capítulos 2–4 já produzida e inserida; não é recalculada como indicador",
        },
        "invariantes": invariantes,
        "integridade": {
            "arquivos_auditados": len(integridade),
            "png_verificados_por_decodificacao_e_dimensao": sum(str(r["arquivo"]).endswith(".png") for r in integridade),
            "svg_verificados_por_xml": sum(str(r["arquivo"]).endswith(".svg") for r in integridade),
            "csv_verificados_por_cabecalho_e_linhas": sum(str(r["arquivo"]).endswith(".csv") for r in integridade),
            "sha256_registrado": True,
        },
        "politica_visual": "QA automatizado verifica cobertura e integridade; inspeção visual editorial final permanece uma checagem humana separada antes da revisão humana do relatório.",
        "referencias_historicas": {
            "M08_p75": 1255,
            "M09_persistentes_p80": 959,
            "M09_mesmo_vetor": 886,
            "regra": "referências de QA; não calibrar a edição corrente; 987/800 permanecem obsoletos",
        },
        "saidas": [
            str(manifesto_csv.relative_to(paths.data_root)),
            str(integridade_csv.relative_to(paths.data_root)),
            str(qa_json.relative_to(paths.data_root)),
        ],
    }
    qa_json.write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    for path in (manifesto_csv, integridade_csv, qa_json):
        registrar_arquivo(manifesto_execucao, path, origem="Etapa 11e — manifesto visual e QA de cobertura")
    registrar_evento(manifesto_execucao, {"tipo": "etapa", "etapa": "11e", "status": qa["status"], "elementos": len(inventario), "pipeline": n_pipeline, "editoriais_externos": n_externos})
    print(json.dumps(qa, ensure_ascii=False, indent=2))
