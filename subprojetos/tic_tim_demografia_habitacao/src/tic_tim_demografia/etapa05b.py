from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

from .fontes.http import HttpClient
from .paths import resolve_paths
from .proveniencia import registrar_evento


VARIAVEIS_AER = [
    "V00001", "V00464", "V00200", "V00201",
    "V00312", "V00313", "V00314", "V00315", "V00316",
    "V00399", "V00400", "V00401", "V00402",
]
VAR_BUEIRO_RE = re.compile(r"\bV0(?:50|52|54)\d{2}\b", re.I)
UNIVERSO_POR_PREFIXO = {"V050": "domicilios", "V052": "moradores", "V054": "faces"}


def _carregar_json(path: Path) -> dict:
    dados = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(dados, dict):
        raise ValueError(f"JSON estrutural invalido: {path}")
    return dados


def _nome(url: str) -> str:
    return Path(urlparse(url).path).name


def _selecionar_domicilios(candidatos: list[str]) -> list[str]:
    """Mantem somente os tres arquivos gerais de caracteristicas domiciliares."""
    escolhidos = []
    for url in candidatos:
        nome = _nome(url).casefold()
        if any(f"caracteristicas_domicilio{i}" in nome for i in (1, 2, 3)):
            escolhidos.append(url)
    if len(escolhidos) != 3:
        raise ValueError(f"Esperados tres arquivos gerais de domicilio; encontrados={escolhidos}")
    return sorted(escolhidos)


def _detectar_encoding(bruto: bytes) -> str:
    for enc in ("utf-8-sig", "cp1252", "latin1"):
        try:
            bruto.decode(enc, errors="strict")
            return enc
        except UnicodeDecodeError:
            continue
    raise ValueError("Codificacao nao reconhecida no cabecalho CSV.")


def _cabecalho_csv(bruto: bytes) -> tuple[str, str, list[str]]:
    enc = _detectar_encoding(bruto)
    texto = bruto.decode(enc, errors="strict")
    primeira = texto.splitlines()[0] if texto.splitlines() else ""
    contagens = {";": primeira.count(";"), ",": primeira.count(","), "\t": primeira.count("\t")}
    sep, n = max(contagens.items(), key=lambda item: item[1])
    if n == 0:
        raise ValueError("Separador CSV nao reconhecido.")
    colunas = next(csv.reader(io.StringIO(primeira), delimiter=sep))
    return enc, sep, [str(c).strip() for c in colunas]


def inspecionar_zip(path: Path) -> dict:
    """Le apenas membros e a primeira linha dos CSVs, sem carregar a base integral."""
    csvs = []
    with zipfile.ZipFile(path) as zf:
        membros = zf.namelist()
        for membro in membros:
            if not membro.casefold().endswith(".csv"):
                continue
            with zf.open(membro) as f:
                bruto = f.readline()
            enc, sep, colunas = _cabecalho_csv(bruto)
            csvs.append(
                {
                    "membro": membro,
                    "encoding": enc,
                    "separador": sep,
                    "n_colunas": len(colunas),
                    "colunas": colunas,
                }
            )
    if not csvs:
        raise ValueError(f"ZIP sem CSV: {path}")
    return {"arquivo": path.name, "membros": membros, "csvs": csvs}


def _baixar_se_ausente(cliente: HttpClient, url: str, destino: Path, manifesto: Path) -> Path:
    if destino.exists():
        return destino
    return cliente.baixar_arquivo(url, destino, manifesto=manifesto)


def _mapear_variaveis(inspecoes: list[dict], variaveis: list[str]) -> dict[str, list[str]]:
    mapa = {v: [] for v in variaveis}
    for item in inspecoes:
        for csv_info in item["csvs"]:
            colunas = set(csv_info["colunas"])
            for var in variaveis:
                if var in colunas:
                    mapa[var].append(f"{item['arquivo']}::{csv_info['membro']}")
    return mapa


def _normalizar_texto(valor: object) -> str:
    return " ".join(str(valor).replace("\n", " ").split()).strip()


def inspecionar_dicionario_bueiro(path: Path) -> list[dict]:
    """Localiza no dicionario oficial linhas que documentam bueiro/boca de lobo.

    O procedimento nao presume o codigo da variavel. Ele procura semanticamente o
    quesito no XLSX e somente depois extrai codigos V050xx/V052xx/V054xx presentes
    na propria linha (ou, se necessario, no contexto imediato de uma linha).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    achados: list[dict] = []
    for ws in wb.worksheets:
        linhas = [
            [_normalizar_texto(v) for v in row]
            for row in ws.iter_rows(values_only=True)
        ]
        for i, valores in enumerate(linhas):
            texto = " | ".join(v for v in valores if v)
            cf = texto.casefold()
            if "bueiro" not in cf and "boca de lobo" not in cf and "boca-de-lobo" not in cf:
                continue
            codigos = sorted(set(VAR_BUEIRO_RE.findall(texto.upper())))
            contexto = texto
            if not codigos:
                inicio = max(0, i - 1)
                fim = min(len(linhas), i + 2)
                contexto = " || ".join(
                    " | ".join(v for v in linhas[j] if v) for j in range(inicio, fim)
                )
                codigos = sorted(set(VAR_BUEIRO_RE.findall(contexto.upper())))
            achados.append(
                {
                    "planilha": ws.title,
                    "linha": i + 1,
                    "texto": texto,
                    "contexto": contexto if contexto != texto else None,
                    "codigos": codigos,
                }
            )
    wb.close()
    return achados


def _categoria_bueiro(texto: str) -> str | None:
    cf = texto.casefold()
    if "nao declarado" in cf or "não declarado" in cf or "não-declarado" in cf:
        return "nao_declarado"
    if "sem bueiro" in cf or "sem boca de lobo" in cf or "sem boca-de-lobo" in cf:
        return "nao"
    if re.search(r"(?:^|\W)(?:nao|não)(?:\W|$)", cf):
        return "nao"
    if "com bueiro" in cf or "com boca de lobo" in cf or "com boca-de-lobo" in cf:
        return "sim"
    if re.search(r"(?:^|\W)sim(?:\W|$)", cf):
        return "sim"
    return None


def resolver_codigos_bueiro(achados: list[dict]) -> dict[str, dict[str, str]]:
    """Resolve codigo por universo/categoria usando exclusivamente evidencia do dicionario."""
    candidatos: dict[str, dict[str, set[str]]] = {
        u: {"sim": set(), "nao": set(), "nao_declarado": set()}
        for u in UNIVERSO_POR_PREFIXO.values()
    }
    for item in achados:
        texto = item.get("texto") or item.get("contexto") or ""
        categoria = _categoria_bueiro(str(texto))
        if categoria is None:
            continue
        for codigo in item.get("codigos", []):
            codigo = str(codigo).upper()
            universo = UNIVERSO_POR_PREFIXO.get(codigo[:4])
            if universo:
                candidatos[universo][categoria].add(codigo)

    resolvido: dict[str, dict[str, str]] = {}
    for universo, cats in candidatos.items():
        if all(len(cats[c]) == 1 for c in ("sim", "nao", "nao_declarado")):
            resolvido[universo] = {c: next(iter(cats[c])) for c in cats}
    return resolvido


def _candidatos_documentacao(*snapshots: dict) -> list[str]:
    links: list[str] = []
    for snap in snapshots:
        links.extend(str(x) for x in snap.get("links", []))
    candidatos = []
    for url in links:
        nome = _nome(url).casefold()
        if "dicion" in nome and nome.endswith(".xlsx"):
            candidatos.append(url)
    return sorted(set(candidatos))


def _validar_codigos_em_cabecalhos(
    codigos: dict[str, dict[str, str]], inspecoes_entorno: dict[str, dict]
) -> bool:
    for universo, categorias in codigos.items():
        colunas = {
            c
            for info in inspecoes_entorno[universo]["csvs"]
            for c in info["colunas"]
        }
        if not all(codigo in colunas for codigo in categorias.values()):
            return False
    return len(codigos) == 3


def executar(raiz: Path) -> None:
    raiz = raiz.resolve()
    paths = resolve_paths(raiz)
    paths.create()
    manifesto = paths.manifests / "execucao.jsonl"

    qa05a_path = paths.qa / "etapa05a_selecao_fontes_isau.json"
    if not qa05a_path.exists():
        raise FileNotFoundError(f"Gate 05a ausente: {qa05a_path}. Execute primeiro --etapa 05a.")
    qa05a = _carregar_json(qa05a_path)

    urls_dom = _selecionar_domicilios(list(qa05a["candidatos_agregados_domiciliares"]))
    entorno = qa05a["entorno"]["candidatos_por_universo"]
    urls_entorno: dict[str, str] = {}
    for universo in ("domicilios", "moradores", "faces"):
        candidatos = list(entorno.get(universo, []))
        if len(candidatos) != 1:
            raise ValueError(f"Entorno {universo}: selecao nao unica: {candidatos}")
        urls_entorno[universo] = candidatos[0]

    cliente = HttpClient(timeout=600)
    raw_dom = paths.raw / "ibge" / "censo2022" / "isau" / "domicilios"
    raw_ent = paths.raw / "ibge" / "censo2022" / "isau" / "entorno"
    raw_doc = paths.raw / "ibge" / "censo2022" / "isau" / "documentacao"
    raw_dom.mkdir(parents=True, exist_ok=True)
    raw_ent.mkdir(parents=True, exist_ok=True)
    raw_doc.mkdir(parents=True, exist_ok=True)

    inspecoes_dom = []
    for url in urls_dom:
        path = _baixar_se_ausente(cliente, url, raw_dom / _nome(url), manifesto)
        inspecoes_dom.append(inspecionar_zip(path))

    inspecoes_entorno = {}
    for universo, url in urls_entorno.items():
        path = _baixar_se_ausente(cliente, url, raw_ent / _nome(url), manifesto)
        inspecoes_entorno[universo] = inspecionar_zip(path)

    mapa_aer = _mapear_variaveis(inspecoes_dom, VARIAVEIS_AER)
    faltantes = sorted(v for v, fontes in mapa_aer.items() if not fontes)
    ambiguas = {v: fontes for v, fontes in mapa_aer.items() if len(fontes) > 1}

    snap_ent = _carregar_json(
        paths.raw / "ibge" / "indices_publicacao" / "censo2022_entorno_setor.json"
    )
    snap_agreg = _carregar_json(
        paths.raw / "ibge" / "indices_publicacao" / "censo2022_agregados_setor.json"
    )
    candidatos_dicionario = _candidatos_documentacao(snap_ent, snap_agreg)

    evidencias_dicionario: list[dict] = []
    codigos_bueiro: dict[str, dict[str, str]] = {}
    dicionario_resolvedor: str | None = None
    erros_dicionario: list[dict] = []
    for url in candidatos_dicionario:
        try:
            path = _baixar_se_ausente(cliente, url, raw_doc / _nome(url), manifesto)
            achados = inspecionar_dicionario_bueiro(path)
            resolvido = resolver_codigos_bueiro(achados)
            evidencias_dicionario.append(
                {
                    "url": url,
                    "arquivo": path.name,
                    "n_linhas_bueiro": len(achados),
                    "achados": achados,
                    "codigos_resolvidos": resolvido,
                }
            )
            if _validar_codigos_em_cabecalhos(resolvido, inspecoes_entorno):
                codigos_bueiro = resolvido
                dicionario_resolvedor = url
                break
        except Exception as exc:  # manter diagnostico auditavel; outro candidato pode resolver
            erros_dicionario.append({"url": url, "erro": f"{type(exc).__name__}: {exc}"})

    status_aer = "RESOLVIDO" if not faltantes and not ambiguas else "PENDENTE"
    status_d = "RESOLVIDO" if codigos_bueiro else "PENDENTE"
    if status_aer == "RESOLVIDO" and status_d == "RESOLVIDO":
        status = "RESOLVIDO_AER_DRENAGEM"
    elif status_aer == "RESOLVIDO":
        status = "RESOLVIDO_AER_PENDENTE_DRENAGEM"
    else:
        status = "DIAGNOSTICO_ESTRUTURAL"

    qa = {
        "status": status,
        "etapa": "05b",
        "arquivos_domiciliares": urls_dom,
        "arquivos_entorno": urls_entorno,
        "inspecao_domicilios": inspecoes_dom,
        "inspecao_entorno": inspecoes_entorno,
        "mapa_variaveis_aer": mapa_aer,
        "variaveis_aer_faltantes": faltantes,
        "variaveis_aer_ambiguas": ambiguas,
        "candidatos_documentacao_entorno": candidatos_dicionario,
        "dicionario_resolvedor_bueiro": dicionario_resolvedor,
        "codigos_bueiro_por_universo": codigos_bueiro,
        "evidencias_dicionario_bueiro": evidencias_dicionario,
        "erros_dicionario": erros_dicionario,
        "regra": (
            "A/E/R somente podem ser calculados depois de todas as variaveis requeridas serem "
            "localizadas em cabecalhos oficiais; D somente depois de os codigos de bueiro/boca de "
            "lobo serem comprovados no dicionario oficial e reencontrados nos cabecalhos dos tres universos"
        ),
        "regra_drenagem": (
            "Em cada universo, Sim e Nao formam o denominador valido; Nao declarado permanece "
            "ausente da proporcao substantiva. O calculo de D sera delegado ao 05c, que le os codigos deste QA."
        ),
        "proximo_gate": (
            "05c calcula A/E/R/D e ISAU C4/C3 somente se status=RESOLVIDO_AER_DRENAGEM"
        ),
    }
    destino = paths.qa / "etapa05b_inspecao_fontes_isau.json"
    destino.write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    registrar_evento(
        manifesto,
        {"tipo": "etapa", "etapa": "05b", "status": qa["status"], "saida": str(destino.relative_to(paths.data_root))},
    )
    print(json.dumps(qa, ensure_ascii=False, indent=2))
